import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Union

import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font
from pandas import DataFrame

from docling_eval.aggregations.multi_evalutor import MultiEvaluation
from docling_eval.datamodels.types import (
    BenchMarkNames,
    ConsolidationFormats,
    EvaluationModality,
)
from docling_eval.evaluators.base_evaluator import EvaluationRejectionType
from docling_eval.evaluators.bbox_text_evaluator import DatasetBoxesTextEvaluation
from docling_eval.evaluators.layout_evaluator import DatasetLayoutEvaluation
from docling_eval.evaluators.markdown_text_evaluator import DatasetMarkdownEvaluation
from docling_eval.evaluators.readingorder_evaluator import DatasetReadingOrderEvaluation
from docling_eval.evaluators.stats import DatasetStatistics
from docling_eval.evaluators.table_evaluator import DatasetTableEvaluation

_log = logging.getLogger(__name__)


def export_value(val: Union[float, DatasetStatistics]) -> str:
    r"""Get statistics value"""
    if isinstance(val, DatasetStatistics):
        fmt_val = f"{val.mean:.2f}±{val.std:.2f}"
    else:
        fmt_val = f"{val:.2f}"

    return fmt_val


class Consolidator:
    r"""
    Consolidate a MultiEvaluation into a comparison matrix

    The comparison matrix has 3 dimensions:
    - Benchmarks
    - ConversionProviders
    - Modalities
    """

    def __init__(self, output_path: Path):
        r""" """
        self._output_path = output_path
        self._excel_engine = "openpyxl"
        self._sheet_name = "matrix"
        self._outfile_stem = "consolidation_matrix"

        self._output_path.mkdir(parents=True, exist_ok=True)

    def __call__(
        self,
        multi_evaluation: MultiEvaluation,
        consolidation_format: Optional[
            ConsolidationFormats
        ] = ConsolidationFormats.EXCEL,
    ) -> Tuple[Dict[EvaluationModality, DataFrame], Optional[List[Path]]]:
        r""" """
        dfs = self._build_dataframes(multi_evaluation)

        # Export dataframe
        if consolidation_format == ConsolidationFormats.EXCEL:
            produced_fn = self._to_excel(dfs)
            _log.info("Produced excel file: %s", str(produced_fn))
            produced_fns = [produced_fn]
        elif consolidation_format == ConsolidationFormats.LATEX:
            benchmark_names: List[BenchMarkNames] = list(
                multi_evaluation.evaluations.keys()
            )
            produced_fns = self._to_latex(benchmark_names, dfs)
            _log.info(
                "Produced latex files: %s", " ".join(str(fn) for fn in produced_fns)
            )
        else:
            _log.info("Unsupported consolidation format: %s", consolidation_format)

        return dfs, produced_fns

    def _to_latex(
        self,
        benchmark_names: List[BenchMarkNames],
        dfs: Dict[EvaluationModality, DataFrame],
    ) -> List[Path]:
        r"""
        Export to LaTeX
        Produce separate tex files for each modality and benchmark
        """
        # Select the dataframe headers and their mapping to latex table headers
        latex_headers: Dict[str, str] = {
            "Experiment": "Experiment",
            "mAP": "mAP",
            "segmentation_f1": "segmentation-f1",
            "weighted_mAP_50": "weighted-mAP-50",
            "weighted_mAP_75": "weighted-mAP-75",
            "weighted_mAP_90": "weighted-mAP-90",
        }

        def align_latex_ampersands(latex_str: str) -> str:
            r"""Format the latex code to have all columns aligned"""
            lines = latex_str.splitlines()
            table_lines = []
            output_lines = []

            for line in lines:
                if "&" in line:
                    parts = [p.strip() for p in line.split("&")]
                    table_lines.append(parts)
                else:
                    # Flush aligned table lines if any
                    if table_lines:
                        # Transpose to get max width per column
                        col_widths = [
                            max(len(row[i]) for row in table_lines)
                            for i in range(len(table_lines[0]))
                        ]
                        for row in table_lines:
                            padded = [
                                val.ljust(col_widths[i]) for i, val in enumerate(row)
                            ]
                            output_lines.append(" & ".join(padded))
                        table_lines = []
                    output_lines.append(line)

            # In case table ends without a non-& line
            if table_lines:
                col_widths = [
                    max(len(row[i]) for row in table_lines)
                    for i in range(len(table_lines[0]))
                ]
                for row in table_lines:
                    padded = [val.ljust(col_widths[i]) for i, val in enumerate(row)]
                    output_lines.append(" & ".join(padded))

            return "\n".join(output_lines)

        # Filter the dataframe
        latexes: List[Path] = []
        for evaluation_modality, df in dfs.items():
            for benchmark in benchmark_names:
                # Select the benchmark rows
                b_df = df[df["Benchmark"] == benchmark]

                # Filter to keep only the latex columns and rename according to the mapping
                b_df = b_df[list(latex_headers.keys())]
                b_df.rename(columns=lambda x: latex_headers[x], inplace=True)
                b_df = b_df.applymap(  # type: ignore
                    lambda x: x.replace("_", "-") if isinstance(x, str) else x
                )
                b_df = b_df.sort_values(
                    by=["Experiment"],
                    ascending=[True],
                )

                # Generate latex
                latex: str = b_df.to_latex(index=False, float_format="%.2f")
                latex = align_latex_ampersands(latex)

                # Save latex
                latex_fn = (
                    self._output_path
                    / f"{self._outfile_stem}_{evaluation_modality.value}_{benchmark.value}.tex"
                )
                with open(latex_fn, "w") as fd:
                    fd.write(latex)
                latexes.append(latex_fn)

        return latexes

    def _to_excel(self, dfs: Dict[EvaluationModality, DataFrame]) -> Path:
        r"""
        Export to Excel
        Produce a single spreadsheet for all modalities
        """
        excel_fn = self._output_path / f"{self._outfile_stem}.xlsx"
        startrow = 0
        header_rows: List[int] = []
        with pd.ExcelWriter(excel_fn, engine=self._excel_engine) as writer:  # type: ignore
            for modality, df in dfs.items():
                if self._sheet_name in writer.book.sheetnames:
                    sheet = writer.book[self._sheet_name]
                    startrow = sheet.max_row + 2

                # Add the modality as a "header" for the metrics subtable
                header_df = DataFrame([modality.name])
                header_rows.append(startrow + 1)
                header_df.to_excel(
                    writer,
                    sheet_name=self._sheet_name,
                    startrow=startrow,
                    index=False,
                    header=False,
                )
                startrow += 1

                # Metrics subtable
                df.to_excel(
                    writer,
                    sheet_name=self._sheet_name,
                    startrow=startrow,
                    index=False,
                )
        # Format the excel
        self._format_excel(excel_fn, header_rows)

        return excel_fn

    def _format_excel(self, excel_fn: Path, header_rows: List[int]):
        r"""Do some proper formatting of the generated excel"""
        workbook = load_workbook(excel_fn)
        sheet = workbook[self._sheet_name]

        # Adjust the cell width
        for col in sheet.columns:
            # Find the maximum length of strings in this column (excluding empty cells)
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2  # Add some padding to make it look better
            first_cell = col[0]
            assert isinstance(first_cell, Cell)
            sheet.column_dimensions[first_cell.column_letter].width = adjusted_width

        # Iterate through each cell in the worksheet and remove borders
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = openpyxl.styles.Border()  # Remove borders

        # Make bold the subtable headers
        bold_font = Font(bold=True)
        for header_row in header_rows:
            cell = sheet.cell(row=header_row, column=1)
            cell.font = bold_font
            x = 0

        # Save back the excel
        workbook.save(excel_fn)

    def _build_dataframes(
        self,
        multi_evaluation: MultiEvaluation,
    ) -> Dict[EvaluationModality, DataFrame]:
        r"""
        Return a Dict with dataframes per modality
        """
        # Collect all data to build the dataframes
        df_data: Dict[EvaluationModality, List[Dict[str, Union[str, float, int]]]] = {}

        # Collect the dataframe data
        for benchmark, exp_mod_eval in multi_evaluation.evaluations.items():
            for experiment, mod_eval in exp_mod_eval.items():
                for modality, single_evaluation in mod_eval.items():
                    evaluation = single_evaluation.evaluation

                    if modality == EvaluationModality.LAYOUT:
                        metrics = self._layout_metrics(evaluation)
                    elif modality == EvaluationModality.MARKDOWN_TEXT:
                        metrics = self._markdowntext_metrics(evaluation)
                    elif modality == EvaluationModality.TABLE_STRUCTURE:
                        metrics = self._tablestructure_metrics(evaluation)
                    elif modality == EvaluationModality.READING_ORDER:
                        metrics = self._readingorder_metrics(evaluation)
                    elif modality == EvaluationModality.BBOXES_TEXT:
                        metrics = self._bboxestext_metrics(evaluation)
                    else:
                        _log.error(
                            "Evaluation modality unsupported for export: %s", modality
                        )
                        continue

                    # Gather the dataframe data
                    data: Dict[str, Union[str, float]] = {
                        "Benchmark": benchmark.value,
                        "Experiment": experiment,
                        "evaluated_samples": evaluation.evaluated_samples,
                    }
                    if single_evaluation.prediction_provider_type is not None:
                        data["Provider"] = (
                            single_evaluation.prediction_provider_type.value
                        )

                    for rej_type in EvaluationRejectionType:
                        if rej_type not in evaluation.rejected_samples:
                            data[rej_type.value] = 0
                        else:
                            data[rej_type.value] = evaluation.rejected_samples[rej_type]

                    data |= metrics
                    if modality not in df_data:
                        df_data[modality] = []
                    df_data[modality].append(data)

        # Build the dataframes
        dfs: Dict[EvaluationModality, DataFrame] = {}
        for modality, m_data in df_data.items():
            df = DataFrame(m_data)
            df = df.sort_values(by=["Benchmark", "Experiment"], ascending=[True, True])
            dfs[modality] = df

        return dfs

    def _layout_metrics(self, evaluation: DatasetLayoutEvaluation) -> Dict[str, str]:
        r"""Get the metrics for the LayoutEvaluation"""
        metrics = {
            "mAP": export_value(evaluation.mAP),
            "stat_mAP": export_value(evaluation.map_stats),
            "stat_mAP_50": export_value(evaluation.map_50_stats),
            "stat_mAP_75": export_value(evaluation.map_75_stats),
            "weighted_mAP_50": export_value(evaluation.weighted_map_50_stats),
            "weighted_mAP_75": export_value(evaluation.weighted_map_75_stats),
            "weighted_mAP_90": export_value(evaluation.weighted_map_90_stats),
            "weighted_mAP_95": export_value(evaluation.weighted_map_95_stats),
            "segmentation_precision": export_value(
                evaluation.segmentation_precision_stats
            ),
            "segmentation_recall": export_value(evaluation.segmentation_recall_stats),
            "segmentation_f1": export_value(evaluation.segmentation_f1_stats),
        }
        for class_evaluation in evaluation.evaluations_per_class:
            key = f"class_{class_evaluation.label}"
            metrics[key] = export_value(class_evaluation.value)

        return metrics

    def _markdowntext_metrics(
        self,
        evaluation: DatasetMarkdownEvaluation,
    ) -> Dict[str, str]:
        r""" """
        metrics = {
            "BLEU": export_value(evaluation.bleu_stats),
            "F1": export_value(evaluation.f1_score_stats),
            "Precision": export_value(evaluation.precision_stats),
            "Recall": export_value(evaluation.recall_stats),
            "Edit_Distance": export_value(evaluation.edit_distance_stats),
            "METEOR": export_value(evaluation.meteor_stats),
        }
        return metrics

    def _tablestructure_metrics(
        self,
        evaluation: DatasetTableEvaluation,
    ) -> Dict[str, str]:
        r""" """
        metrics = {
            "TEDS": export_value(evaluation.TEDS),
            "TEDS_struct": export_value(evaluation.TEDS_struct),
            "TEDS_simple": export_value(evaluation.TEDS_simple),
            "TEDS_complex": export_value(evaluation.TEDS_complex),
        }
        return metrics

    def _readingorder_metrics(
        self,
        evaluation: DatasetReadingOrderEvaluation,
    ) -> Dict[str, str]:
        r""" """
        metrics = {
            "ARD": export_value(evaluation.ard_stats),
            "Weighted_ARD": export_value(evaluation.w_ard_stats),
        }
        return metrics

    def _bboxestext_metrics(
        self,
        evaluation: DatasetBoxesTextEvaluation,
    ) -> Dict[str, str]:
        r""" """
        metrics = {
            "BLEU": export_value(evaluation.bleu_stats),
            "F1": export_value(evaluation.f1_score_stats),
            "Precision": export_value(evaluation.precision_stats),
            "Recall": export_value(evaluation.recall_stats),
            "Edit_Distance": export_value(evaluation.edit_distance_stats),
            "METEOR": export_value(evaluation.meteor_stats),
        }
        return metrics
